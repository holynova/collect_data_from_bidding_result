# -*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook
import os,datetime
import logging
logging.basicConfig(level = logging.INFO)


folder = r"E:\kuaipan\github\collect_data_from_bidding_result\bid"
excel_files =[]
class Bid():
	pass
class Pack():
	pass
for root, dirs, files in os.walk(folder, topdown=True):
	cnt_file = 0
	cnt_all_bid = 0
	cnt_all_pack = 0
	cnt_all_other = 0
	for f in files:
		if f.find('.xlsx') != -1:
			wb_name = os.path.join(root,f)
			try:
				wb = load_workbook(wb_name)
			except:
				print 'Error : fail to open ',wb_name
				continue
			shts = wb.get_sheet_names()
			
			print cnt_file,':',f
			cnt_file += 1

			cnt_bid = 0
			cnt_pack = 0
			cnt_other = 0
			if ' '.join(shts).find(u"基本信息") != -1:
				for sht_name in shts:
					sht = wb.get_sheet_by_name(sht_name)
					if sht_name.find(u'基本') != -1 or sht['a1'].value == u'投标信息组':
						#collect basic info
						bid = Bid()
						bid.work_num = sht['b2'].value
						bid.date = sht['b3'].value
						bid.sr = sht['b4'].value
						bid.csc = sht['b5'].value
						bid.scoring = sht['b6'].value
						bid.sheets = []
						cnt_bid += 1
						print '--',bid.work_num

					elif sht_name.find(u'包') != -1 or (sht['a1'].value == u'序号' and sht['b1'].value == u'厂家'):
						#collect pack info
						pack = Pack()
						sht['s1'].value = u"包号"
						sht['s2'].value = u"nkt_gm3"
						sht['s3'].value = u"投标厂家数"
						sht['s4'].value = u"中标厂家"
						sht['s5'].value = u"最高价"
						sht['s6'].value = u"最低价"
						sht['s7'].value = u"平均价"
						sht['s8'].value = u"去掉极值后的平均价"
						sht['s9'].value = u"中位数"
						sht['s10'].value = u"中标价"
						sht['s11'].value = u"安凯特价格"

						sht['t1'].value = '=I2'
						sht['t2'].value = '=I7'
						sht['t3'].value = '=COUNTIF(C:C,">0")'
						sht['t4'].value = '=INDEX(B:B,MATCH("中标",D:D,0))'
						sht['t5'].value = '=MAX(C:C)'
						sht['t6'].value = '=MIN(C:C)'
						sht['t7'].value = '=AVERAGE(C:C)'
						sht['t8'].value = '=TRIMMEAN(C:C,0.04)'
						sht['t9'].value = '=MEDIAN(C:C)'
						sht['t10'].value = '=INDEX(C:C,MATCH("中标",D:D,0))'
						sht['t11'].value = '=INDEX(C:C,MATCH("NKT",B:B,0))'
						cnt_pack += 1
					else:
						cnt_other += 1
				print "--finished with:\n--%d bids\n--%d packs\n--%d other " %(cnt_bid,cnt_pack,cnt_other)
				cnt_all_other += cnt_other
				cnt_all_pack += cnt_pack
				cnt_all_bid += cnt_bid
			wb.save(wb_name)		
print 'all\n%d bids\n%d packs\n%d others' %(cnt_all_bid,cnt_all_pack,cnt_all_other)
			# else :
			# 	print 'sht = %s\n|--a1 = %s\n|--a2 = %s' %(sht_name,sht['a1'].value,sht['b1'].value)
				# print d 

    # for name in files:
    #     # print( 'files----'+os.path.join(root, name))
    #     print os.path.join(root,name)
    #     excel_files.append(os.path.join(root, name))
# i=0
# for f in excel_files:
# 	print "%d:%s" %(i,f)
# 	i += 1 

# for f in files:
# 	wb = load_workbook(f)
# 	for sht_name in wb.get_sheet_names():
# 		print sht_name
		# if sht_name.find(u'基本信息') != -1 or sht_name['a1'] == u'投标信息组':
		# 	#collect basic info
		# 	print '%s in %s' %(sht_name,f)
		# elif sht_name.find(u'包') != -1 or (sht_name['a1'] == u'序号' and sht_name['a2'] == u'厂家'):
		# 	#collect pack info
		# 	print '\t %s in %s' %(sht_name,f)
print 'done!'


# import os


# import os

# for f in folder:
# 	wb = xlrd.open_workbook(f)
# 	for sht in wb:
# 		if sht.name = "基本信息" or sht.cell("a1") == "投标信息组":
# 			bidInfo = [name,date,SR,company,workNum]
# 			# pass

# 		elif sht.name = "包xx" or (sht.cell('a1') == "序号" and sht.cell('a2') == "厂家"):
# 			包号  i2
# 			gm3  i7
# 			nkt价格
# 			中标价
# 			中标厂家
# 			价格数组


# plan B
# 先用openpyxl打开文件,把每个package页面用excel的公式处理好,然后再统一抓取

